# MedCase Pro Platform - Import Servisi
# Excel va CSV dan ma'lumotlarni import qilish

from typing import List, Dict, Any, Tuple, Optional
from uuid import uuid4, UUID
from datetime import datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
import io
import re
import unicodedata

# Excel uchun
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_MAVJUD = True
except ImportError:
    EXCEL_MAVJUD = False

from modellar.kategoriya import AsosiyKategoriya, KichikKategoriya, Bolim
from modellar.holat import Holat, HolatVarianti, HolatMedia, QiyinlikDarajasi, HolatTuri, MediaTuri


def slugify(text: str) -> str:
    """Matnni slug formatiga o'zgartirish."""
    if not text:
        return ""
    # Kichik harfga o'tkazish
    text = text.lower().strip()
    # Unicode belgilarni ASCII ga o'zgartirish
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    # Bo'sh joylar va maxsus belgilarni tire bilan almashtirish
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '-', text)
    return text.strip('-')


def normalize_text(text: str) -> str:
    """Matnni normallashtirish - kichik harf, bo'sh joylarni tozalash."""
    if not text:
        return ""
    return text.lower().strip()


def title_case(text: str) -> str:
    """Matnni Title Case formatiga o'zgartirish."""
    if not text:
        return ""
    return text.strip().title()


class ImportServisi:
    """
    Excel va CSV dan ma'lumotlarni import qilish servisi.
    """
    
    def __init__(self, db: AsyncSession):
        self.db = db
        self.xatolar: List[Dict] = []
        self.ogohlantirishlar: List[Dict] = []
    
    # Qiyinlik darajasini mapping
    QIYINLIK_MAP = {
        'basic': QiyinlikDarajasi.OSON,
        'oson': QiyinlikDarajasi.OSON,
        'easy': QiyinlikDarajasi.OSON,
        'intermediate': QiyinlikDarajasi.ORTACHA,
        'ortacha': QiyinlikDarajasi.ORTACHA,
        'orta': QiyinlikDarajasi.ORTACHA,
        'medium': QiyinlikDarajasi.ORTACHA,
        'advanced': QiyinlikDarajasi.QIYIN,
        'qiyin': QiyinlikDarajasi.QIYIN,
        'hard': QiyinlikDarajasi.QIYIN,
    }
    
    async def excel_tahlil_qilish(self, fayl_content: bytes) -> Dict[str, Any]:
        """
        Excel faylni tahlil qiladi va xatoliklarni tekshiradi.
        Import qilmasdan oldin faylni validatsiya qilish uchun.
        """
        if not EXCEL_MAVJUD:
            return {
                "muvaffaqiyat": False,
                "xato": "openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl"
            }
        
        self.xatolar = []
        self.ogohlantirishlar = []
        
        try:
            # Excel faylni o'qish
            workbook = load_workbook(io.BytesIO(fayl_content), data_only=True)
            sheet = workbook.active
            
            # Ustun nomlarini olish (birinchi qator)
            ustunlar = {}
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    ustunlar[cell.value.lower().strip().replace('-', '_')] = col_idx
            
            # Kerakli ustunlarni tekshirish (id va diff ixtiyoriy)
            kerakli_ustunlar = [
                'section', 'main_category', 'sub_category',
                'case', 'question', 'opt_a', 'opt_b', 'opt_c', 'opt_d',
                'correct'
            ]
            
            yetishmagan = []
            for ustun in kerakli_ustunlar:
                if ustun not in ustunlar:
                    yetishmagan.append(ustun)
            
            if yetishmagan:
                return {
                    "muvaffaqiyat": False,
                    "xato": f"Kerakli ustunlar topilmadi: {', '.join(yetishmagan)}",
                    "mavjud_ustunlar": list(ustunlar.keys())
                }
            
            # Qatorlarni tahlil qilish
            holatlar = []
            kategoriyalar = set()
            kichik_kategoriyalar = set()
            bolimlar = set()
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                # Bo'sh qatorni o'tkazib yuborish
                if not any(cell.value for cell in row):
                    continue
                
                def get_cell(ustun_nomi):
                    col_idx = ustunlar.get(ustun_nomi)
                    if col_idx and col_idx <= len(row):
                        return row[col_idx - 1].value
                    return None
                
                # Ma'lumotlarni olish
                holat_id = get_cell('id')
                main_category = get_cell('main_category')
                sub_category = get_cell('sub_category')
                section = get_cell('section')
                case_text = get_cell('case')
                question = get_cell('question')
                opt_a = get_cell('opt_a')
                opt_b = get_cell('opt_b')
                opt_c = get_cell('opt_c')
                opt_d = get_cell('opt_d')
                correct = get_cell('correct')
                expl_a = get_cell('expl_a')
                expl_b = get_cell('expl_b')
                expl_c = get_cell('expl_c')
                expl_d = get_cell('expl_d')
                diff = get_cell('diff')
                link = get_cell('link')
                
                # Validatsiya
                qator_xatolari = []
                
                if not main_category:
                    qator_xatolari.append("main_category bo'sh")
                if not sub_category:
                    qator_xatolari.append("sub_category bo'sh")
                if not section:
                    qator_xatolari.append("section bo'sh")
                if not case_text:
                    qator_xatolari.append("case bo'sh")
                if not question:
                    qator_xatolari.append("question bo'sh")
                if not opt_a:
                    qator_xatolari.append("opt_a bo'sh")
                if not opt_b:
                    qator_xatolari.append("opt_b bo'sh")
                if not opt_c:
                    qator_xatolari.append("opt_c bo'sh")
                if not opt_d:
                    qator_xatolari.append("opt_d bo'sh")
                
                # To'g'ri javobni tekshirish
                if correct:
                    correct = str(correct).upper().strip()
                    if correct not in ['A', 'B', 'C', 'D']:
                        qator_xatolari.append(f"correct noto'g'ri: '{correct}' (A, B, C, D bo'lishi kerak)")
                else:
                    qator_xatolari.append("correct bo'sh")
                
                # Qiyinlik darajasini tekshirish (string sifatida saqlash - JSON uchun)
                qiyinlik = "ortacha"
                if diff:
                    diff_lower = str(diff).lower().strip()
                    diff_key = re.sub(r"[^a-z]", "", diff_lower)
                    if diff_key in self.QIYINLIK_MAP:
                        qiyinlik = self.QIYINLIK_MAP[diff_key].value  # Enum value (string)
                    else:
                        self.ogohlantirishlar.append({
                            "qator": row_idx,
                            "xabar": f"Noma'lum qiyinlik: '{diff}', 'ortacha' sifatida belgilandi"
                        })
                
                if qator_xatolari:
                    self.xatolar.append({
                        "qator": row_idx,
                        "id": holat_id,
                        "xatolar": qator_xatolari
                    })
                else:
                    # Kategoriyalarni to'plash
                    if main_category:
                        kategoriyalar.add(str(main_category).strip().lower())
                    if sub_category and main_category:
                        kichik_kategoriyalar.add((
                            str(main_category).strip().lower(),
                            str(sub_category).strip().lower()
                        ))
                    if section and sub_category:
                        bolimlar.add((
                            str(main_category).strip().lower(),
                            str(sub_category).strip().lower(),
                            str(section).strip().lower()
                        ))
                    
                    holatlar.append({
                        "qator": row_idx,
                        "id": holat_id,
                        "main_category": str(main_category).strip() if main_category else None,
                        "sub_category": str(sub_category).strip() if sub_category else None,
                        "section": str(section).strip() if section else None,
                        "case": str(case_text).strip() if case_text else None,
                        "question": str(question).strip() if question else None,
                        "opt_a": str(opt_a).strip() if opt_a else None,
                        "opt_b": str(opt_b).strip() if opt_b else None,
                        "opt_c": str(opt_c).strip() if opt_c else None,
                        "opt_d": str(opt_d).strip() if opt_d else None,
                        "correct": correct,
                        "expl_a": str(expl_a).strip() if expl_a else None,
                        "expl_b": str(expl_b).strip() if expl_b else None,
                        "expl_c": str(expl_c).strip() if expl_c else None,
                        "expl_d": str(expl_d).strip() if expl_d else None,
                        "qiyinlik": qiyinlik,
                        "link": str(link).strip() if link else None
                    })
            
            return {
                "muvaffaqiyat": len(self.xatolar) == 0,
                "jami_qatorlar": len(holatlar) + len(self.xatolar),
                "yaroqli_holatlar": len(holatlar),
                "xatoli_qatorlar": len(self.xatolar),
                "xatolar": self.xatolar,
                "ogohlantirishlar": self.ogohlantirishlar,
                "kategoriyalar": list(kategoriyalar),
                "kichik_kategoriyalar": len(kichik_kategoriyalar),
                "bolimlar": len(bolimlar),
                "holatlar": holatlar  # Import uchun tayyorlangan ma'lumotlar
            }
            
        except Exception as e:
            return {
                "muvaffaqiyat": False,
                "xato": f"Excel faylni o'qishda xatolik: {str(e)}"
            }
    
    async def excel_import_qilish(self, holatlar: List[Dict]) -> Dict[str, Any]:
        """
        Tahlil qilingan holatlarni bazaga import qiladi.
        Optimallashtirilgan: batch insert va kamroq flush.
        """
        yaratilgan = 0
        yangilangan = 0
        xatolar = []
        yaratilgan_idlar = []  # Yaratilgan holat ID larini saqlash
        
        # Kategoriyalar keshi
        kategoriya_kesh: Dict[str, UUID] = {}
        kichik_kat_kesh: Dict[Tuple[str, str], UUID] = {}
        bolim_kesh: Dict[Tuple[str, str, str], UUID] = {}
        
        # Avval barcha kategoriyalarni bazadan yuklash (bir marta so'rov)
        try:
            # Asosiy kategoriyalarni yuklash
            result = await self.db.execute(select(AsosiyKategoriya))
            for kat in result.scalars().all():
                kategoriya_kesh[normalize_text(kat.nomi)] = kat.id
            
            # Kichik kategoriyalarni yuklash
            result = await self.db.execute(select(KichikKategoriya))
            for kkat in result.scalars().all():
                # Asosiy kategoriya nomini topish
                for main_key, main_id in kategoriya_kesh.items():
                    if main_id == kkat.asosiy_kategoriya_id:
                        kichik_kat_kesh[(main_key, normalize_text(kkat.nomi))] = kkat.id
                        break
            
            # Bo'limlarni yuklash
            result = await self.db.execute(select(Bolim))
            for bolim in result.scalars().all():
                # Kichik kategoriya kalitini topish
                for (main_key, sub_key), sub_id in kichik_kat_kesh.items():
                    if sub_id == bolim.kichik_kategoriya_id:
                        bolim_kesh[(main_key, sub_key, normalize_text(bolim.nomi))] = bolim.id
                        break
        except Exception as e:
            pass  # Kesh yuklash xatosi bo'lsa, davom etamiz
        
        # Holatlarni batch bo'yicha qayta ishlash
        BATCH_SIZE = 50
        pending_holatlar = []  # Flush kutayotgan holatlar
        
        for idx, holat_data in enumerate(holatlar):
            try:
                # Ma'lumotlarni normallashtirish
                main_cat_raw = holat_data['main_category']
                sub_cat_raw = holat_data['sub_category']
                section_raw = holat_data['section']
                
                main_cat_key = normalize_text(main_cat_raw)
                sub_cat_key_str = normalize_text(sub_cat_raw)
                section_key_str = normalize_text(section_raw)
                
                # 1. Asosiy kategoriyani topish/yaratish
                if main_cat_key not in kategoriya_kesh:
                    main_cat_slug = slugify(main_cat_raw)
                    result = await self.db.execute(
                        select(AsosiyKategoriya).where(
                            (func.lower(AsosiyKategoriya.nomi) == main_cat_key) |
                            (AsosiyKategoriya.slug == main_cat_slug)
                        )
                    )
                    kategoriya = result.scalar_one_or_none()
                    
                    if not kategoriya:
                        kategoriya = AsosiyKategoriya(
                            nomi=title_case(main_cat_raw),
                            slug=main_cat_slug,
                            tavsif=f"{title_case(main_cat_raw)} kategoriyasi",
                            rang="#3B82F6",
                            tartib=0
                        )
                        self.db.add(kategoriya)
                        await self.db.flush()
                    
                    kategoriya_kesh[main_cat_key] = kategoriya.id
                
                kategoriya_id = kategoriya_kesh[main_cat_key]
                
                # 2. Kichik kategoriyani topish/yaratish
                sub_cat_key = (main_cat_key, sub_cat_key_str)
                
                if sub_cat_key not in kichik_kat_kesh:
                    sub_cat_slug = slugify(sub_cat_raw)
                    result = await self.db.execute(
                        select(KichikKategoriya).where(
                            KichikKategoriya.asosiy_kategoriya_id == kategoriya_id,
                            (func.lower(KichikKategoriya.nomi) == sub_cat_key_str) |
                            (KichikKategoriya.slug == sub_cat_slug)
                        )
                    )
                    kichik_kat = result.scalar_one_or_none()
                    
                    if not kichik_kat:
                        kichik_kat = KichikKategoriya(
                            asosiy_kategoriya_id=kategoriya_id,
                            nomi=title_case(sub_cat_raw),
                            slug=sub_cat_slug,
                            tavsif=f"{title_case(sub_cat_raw)} bo'limi",
                            tartib=0
                        )
                        self.db.add(kichik_kat)
                        await self.db.flush()
                    
                    kichik_kat_kesh[sub_cat_key] = kichik_kat.id
                
                kichik_kat_id = kichik_kat_kesh[sub_cat_key]
                
                # 3. Bo'limni topish/yaratish
                bolim_key = (main_cat_key, sub_cat_key_str, section_key_str)
                
                if bolim_key not in bolim_kesh:
                    section_slug = slugify(section_raw)
                    result = await self.db.execute(
                        select(Bolim).where(
                            Bolim.kichik_kategoriya_id == kichik_kat_id,
                            (func.lower(Bolim.nomi) == section_key_str) |
                            (Bolim.slug == section_slug)
                        )
                    )
                    bolim = result.scalar_one_or_none()
                    
                    if not bolim:
                        bolim = Bolim(
                            kichik_kategoriya_id=kichik_kat_id,
                            nomi=title_case(section_raw),
                            slug=section_slug,
                            tavsif=f"{title_case(section_raw)} mavzusi",
                            tartib=0
                        )
                        self.db.add(bolim)
                        await self.db.flush()
                    
                    bolim_kesh[bolim_key] = bolim.id
                
                bolim_id = bolim_kesh[bolim_key]
                
                # 4. Holatni yaratish (yangilashni o'tkazib yuboramiz - tezroq)
                sarlavha = holat_data['question'][:100] if len(holat_data['question']) > 100 else holat_data['question']
                if sarlavha:
                    sarlavha = sarlavha[0].upper() + sarlavha[1:] if len(sarlavha) > 1 else sarlavha.upper()
                
                # Qiyinlik darajasini to'g'ri formatga o'tkazish
                qiyinlik = holat_data.get('qiyinlik', 'ortacha')
                if isinstance(qiyinlik, str):
                    qiyinlik = self.QIYINLIK_MAP.get(qiyinlik.lower(), QiyinlikDarajasi.ORTACHA)
                elif not isinstance(qiyinlik, QiyinlikDarajasi):
                    qiyinlik = QiyinlikDarajasi.ORTACHA
                
                # Ball hisoblash
                if qiyinlik == QiyinlikDarajasi.OSON:
                    ball = 10
                elif qiyinlik == QiyinlikDarajasi.QIYIN:
                    ball = 30
                else:
                    ball = 20
                
                holat = Holat(
                    bolim_id=bolim_id,
                    sarlavha=sarlavha,
                    klinik_stsenariy=holat_data['case'],
                    savol=holat_data['question'],
                    togri_javob=holat_data['correct'],
                    qiyinlik=qiyinlik,
                    turi=HolatTuri.MCQ,
                    ball=ball,
                    chop_etilgan=True,
                    tekshirilgan=True
                )
                self.db.add(holat)
                pending_holatlar.append((holat, holat_data))
                
                # Batch flush - har BATCH_SIZE holatdan keyin
                if len(pending_holatlar) >= BATCH_SIZE:
                    await self.db.flush()
                    
                    # Variantlar va media qo'shish
                    for h, h_data in pending_holatlar:
                        # Variantlar
                        for belgi, matn, tushuntirish in [
                            ('A', h_data['opt_a'], h_data.get('expl_a')),
                            ('B', h_data['opt_b'], h_data.get('expl_b')),
                            ('C', h_data['opt_c'], h_data.get('expl_c')),
                            ('D', h_data['opt_d'], h_data.get('expl_d')),
                        ]:
                            variant = HolatVarianti(
                                holat_id=h.id,
                                belgi=belgi,
                                matn=matn,
                                tushuntirish=tushuntirish,
                                togri=(belgi == h_data['correct'])
                            )
                            self.db.add(variant)
                        
                        # Media (default: RASM)
                        if h_data.get('link'):
                            link = h_data['link']
                            media_turi = MediaTuri.RASM  # Default rasm
                            if any(ext in link.lower() for ext in ['.mp4', '.mov', '.avi', '.webm', '.mkv']):
                                media_turi = MediaTuri.VIDEO
                            
                            # nom maydonini string formatga o'tkazish
                            media_nom = h_data.get('id')
                            if media_nom is not None:
                                media_nom = str(media_nom)
                            else:
                                media_nom = 'media'
                            
                            media = HolatMedia(
                                holat_id=h.id,
                                turi=media_turi,
                                url=link,
                                nom=media_nom,
                                tartib=0
                            )
                            self.db.add(media)
                        
                        yaratilgan_idlar.append(h.id)
                        yaratilgan += 1
                    
                    pending_holatlar = []
                
            except Exception as e:
                xatolar.append({
                    "qator": holat_data.get('qator'),
                    "id": holat_data.get('id'),
                    "xato": str(e)
                })
        
        # Qolgan holatlarni flush qilish
        if pending_holatlar:
            try:
                await self.db.flush()
                
                for h, h_data in pending_holatlar:
                    # Variantlar
                    for belgi, matn, tushuntirish in [
                        ('A', h_data['opt_a'], h_data.get('expl_a')),
                        ('B', h_data['opt_b'], h_data.get('expl_b')),
                        ('C', h_data['opt_c'], h_data.get('expl_c')),
                        ('D', h_data['opt_d'], h_data.get('expl_d')),
                    ]:
                        variant = HolatVarianti(
                            holat_id=h.id,
                            belgi=belgi,
                            matn=matn,
                            tushuntirish=tushuntirish,
                            togri=(belgi == h_data['correct'])
                        )
                        self.db.add(variant)
                    
                    # Media (default: RASM)
                    if h_data.get('link'):
                        link = h_data['link']
                        media_turi = MediaTuri.RASM  # Default rasm
                        if any(ext in link.lower() for ext in ['.mp4', '.mov', '.avi', '.webm', '.mkv']):
                            media_turi = MediaTuri.VIDEO
                        
                        # nom maydonini string formatga o'tkazish
                        media_nom = h_data.get('id')
                        if media_nom is not None:
                            media_nom = str(media_nom)
                        else:
                            media_nom = 'media'
                        
                        media = HolatMedia(
                            holat_id=h.id,
                            turi=media_turi,
                            url=link,
                            nom=media_nom,
                            tartib=0
                        )
                        self.db.add(media)
                    
                    yaratilgan_idlar.append(h.id)
                    yaratilgan += 1
            except Exception as e:
                xatolar.append({"xato": f"Qolgan holatlarni saqlashda xato: {str(e)}"})
        
        # Oxirgi commit
        try:
            await self.db.commit()
        except Exception as e:
            await self.db.rollback()
            return {
                "muvaffaqiyat": False,
                "yaratilgan": yaratilgan,
                "yangilangan": yangilangan,
                "xatolar": xatolar + [{"xato": f"Oxirgi commit xatosi: {str(e)}"}]
            }
        
        # Kategoriyalar statistikasini yangilash
        try:
            await self._kategoriya_statistikalarini_yangilash()
        except Exception as e:
            xatolar.append({"xato": f"Statistika yangilashda xato: {str(e)}"})
        
        return {
            "muvaffaqiyat": len(xatolar) == 0,
            "yaratilgan": yaratilgan,
            "yangilangan": yangilangan,
            "xatolar": xatolar
        }
    
    async def _kategoriya_statistikalarini_yangilash(self) -> None:
        """Import tugagandan keyin kategoriyalar statistikasini yangilaydi."""
        # Bo'limlar holatlar sonini yangilash
        bolimlar_sorov = """
            UPDATE bolimlar SET 
                holatlar_soni = (
                    SELECT COUNT(*) FROM holatlar 
                    WHERE holatlar.bolim_id = bolimlar.id 
                    AND holatlar.faol = true 
                    AND holatlar.chop_etilgan = true
                ),
                oson_holatlar = (
                    SELECT COUNT(*) FROM holatlar 
                    WHERE holatlar.bolim_id = bolimlar.id 
                    AND holatlar.faol = true 
                    AND holatlar.chop_etilgan = true
                    AND lower(holatlar.qiyinlik::text) = 'oson'
                ),
                ortacha_holatlar = (
                    SELECT COUNT(*) FROM holatlar 
                    WHERE holatlar.bolim_id = bolimlar.id 
                    AND holatlar.faol = true 
                    AND holatlar.chop_etilgan = true
                    AND lower(holatlar.qiyinlik::text) = 'ortacha'
                ),
                qiyin_holatlar = (
                    SELECT COUNT(*) FROM holatlar 
                    WHERE holatlar.bolim_id = bolimlar.id 
                    AND holatlar.faol = true 
                    AND holatlar.chop_etilgan = true
                    AND lower(holatlar.qiyinlik::text) = 'qiyin'
                )
        """
        await self.db.execute(text(bolimlar_sorov))
        
        # Kichik kategoriyalar holatlar sonini yangilash
        kichik_kat_sorov = """
            UPDATE kichik_kategoriyalar SET 
                holatlar_soni = (
                    SELECT COALESCE(SUM(bolimlar.holatlar_soni), 0) 
                    FROM bolimlar 
                    WHERE bolimlar.kichik_kategoriya_id = kichik_kategoriyalar.id
                    AND bolimlar.faol = true
                )
        """
        await self.db.execute(text(kichik_kat_sorov))
        
        # Asosiy kategoriyalar holatlar sonini yangilash
        asosiy_kat_sorov = """
            UPDATE asosiy_kategoriyalar SET 
                holatlar_soni = (
                    SELECT COALESCE(SUM(kichik_kategoriyalar.holatlar_soni), 0) 
                    FROM kichik_kategoriyalar 
                    WHERE kichik_kategoriyalar.asosiy_kategoriya_id = asosiy_kategoriyalar.id
                    AND kichik_kategoriyalar.faol = true
                )
        """
        await self.db.execute(text(asosiy_kat_sorov))
        
        await self.db.commit()

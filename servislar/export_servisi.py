# MedCase Pro Platform - Export Servisi
# PDF va Excel export funksiyalari

from typing import List, Dict, Any, Optional
from uuid import UUID
from datetime import datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
import io
import json

# PDF uchun (shartli import)
PDF_MAVJUD = False
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_MAVJUD = True
except ImportError:
    pass

# Excel uchun (shartli import)
EXCEL_MAVJUD = False
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_MAVJUD = True
except ImportError:
    pass

from modellar.rivojlanish import FoydalanuvchiRivojlanishi, HolatUrinishi
from modellar.foydalanuvchi import Foydalanuvchi


class ExportServisi:
    """PDF va Excel export servisi."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def rivojlanish_pdf(
        self,
        foydalanuvchi_id: UUID,
        boshlangich_sana: datetime = None,
        tugash_sana: datetime = None
    ) -> bytes:
        """Foydalanuvchi rivojlanish hisobotini PDF formatda yaratadi."""
        
        if not PDF_MAVJUD:
            raise Exception("reportlab kutubxonasi o'rnatilmagan. pip install reportlab")
        
        # Ma'lumotlarni olish
        malumotlar = await self._rivojlanish_malumotlari_olish(
            foydalanuvchi_id, boshlangich_sana, tugash_sana
        )
        
        # PDF yaratish
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Sarlavha
        sarlavha_style = ParagraphStyle(
            'Sarlavha',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20
        )
        story.append(Paragraph("MedCase Pro - Rivojlanish Hisoboti", sarlavha_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Foydalanuvchi ma'lumotlari
        foyd = malumotlar.get('foydalanuvchi', {})
        info_text = f"""
        <b>Foydalanuvchi:</b> {foyd.get('toliq_ism', 'N/A')}<br/>
        <b>Hisobot sanasi:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>
        <b>Davr:</b> {malumotlar.get('davr', 'Barcha vaqt')}
        """
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 1*cm))
        
        # Umumiy statistika
        story.append(Paragraph("<b>Umumiy Statistika</b>", styles['Heading2']))
        
        statistika = malumotlar.get('statistika', {})
        stat_data = [
            ['Ko\'rsatkich', 'Qiymat'],
            ['Jami yechilgan', str(statistika.get('jami_yechilgan', 0))],
            ['To\'g\'ri javoblar', str(statistika.get('togri_javoblar', 0))],
            ['Aniqlik', f"{statistika.get('aniqlik', 0):.1f}%"],
            ['Joriy streak', f"{statistika.get('joriy_streak', 0)} kun"],
            ['Maksimal streak', f"{statistika.get('maksimal_streak', 0)} kun"],
            ['Jami ball', str(statistika.get('jami_ball', 0))],
            ['Daraja', str(statistika.get('daraja', 1))],
        ]
        
        stat_table = Table(stat_data, colWidths=[8*cm, 4*cm])
        stat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F1F5F9')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 1*cm))
        
        # Kategoriya bo'yicha statistika
        if malumotlar.get('kategoriya_statistika'):
            story.append(Paragraph("<b>Kategoriya Bo'yicha</b>", styles['Heading2']))
            
            kat_data = [['Kategoriya', 'Yechilgan', 'To\'g\'ri', 'Aniqlik']]
            for kat in malumotlar['kategoriya_statistika']:
                kat_data.append([
                    kat['nomi'],
                    str(kat['yechilgan']),
                    str(kat['togri']),
                    f"{kat['aniqlik']:.1f}%"
                ])
            
            kat_table = Table(kat_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
            kat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ]))
            story.append(kat_table)
        
        # PDF yaratish
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    async def rivojlanish_excel(
        self,
        foydalanuvchi_id: UUID,
        boshlangich_sana: datetime = None,
        tugash_sana: datetime = None
    ) -> bytes:
        """Foydalanuvchi rivojlanish hisobotini Excel formatda yaratadi."""
        
        if not EXCEL_MAVJUD:
            raise Exception("openpyxl kutubxonasi o'rnatilmagan")
        
        malumotlar = await self._rivojlanish_malumotlari_olish(
            foydalanuvchi_id, boshlangich_sana, tugash_sana
        )
        
        # Excel yaratish
        wb = openpyxl.Workbook()
        
        # ============== Umumiy sahifa ==============
        ws = wb.active
        ws.title = "Umumiy"
        
        # Sarlavha
        ws['A1'] = "MedCase Pro - Rivojlanish Hisoboti"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        foyd = malumotlar.get('foydalanuvchi', {})
        ws['A3'] = f"Foydalanuvchi: {foyd.get('toliq_ism', 'N/A')}"
        ws['A4'] = f"Hisobot sanasi: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Statistika jadvali
        statistika = malumotlar.get('statistika', {})
        headers = ['Ko\'rsatkich', 'Qiymat']
        data = [
            ['Jami yechilgan', statistika.get('jami_yechilgan', 0)],
            ['To\'g\'ri javoblar', statistika.get('togri_javoblar', 0)],
            ['Aniqlik (%)', round(statistika.get('aniqlik', 0), 1)],
            ['Joriy streak (kun)', statistika.get('joriy_streak', 0)],
            ['Maksimal streak (kun)', statistika.get('maksimal_streak', 0)],
            ['Jami ball', statistika.get('jami_ball', 0)],
            ['Daraja', statistika.get('daraja', 1)],
        ]
        
        # Header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_idx, row_data in enumerate(data, 7):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center")
        
        # ============== Kategoriya sahifasi ==============
        if malumotlar.get('kategoriya_statistika'):
            ws2 = wb.create_sheet("Kategoriya bo'yicha")
            
            kat_headers = ['Kategoriya', 'Yechilgan', 'To\'g\'ri', 'Aniqlik (%)']
            for col, header in enumerate(kat_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            
            for row_idx, kat in enumerate(malumotlar['kategoriya_statistika'], 2):
                ws2.cell(row=row_idx, column=1, value=kat['nomi'])
                ws2.cell(row=row_idx, column=2, value=kat['yechilgan'])
                ws2.cell(row=row_idx, column=3, value=kat['togri'])
                ws2.cell(row=row_idx, column=4, value=round(kat['aniqlik'], 1))
        
        # ============== Kunlik statistika ==============
        if malumotlar.get('kunlik_statistika'):
            ws3 = wb.create_sheet("Kunlik")
            
            kun_headers = ['Sana', 'Yechilgan', 'To\'g\'ri', 'Aniqlik (%)']
            for col, header in enumerate(kun_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
            
            for row_idx, kun in enumerate(malumotlar['kunlik_statistika'], 2):
                ws3.cell(row=row_idx, column=1, value=kun['sana'])
                ws3.cell(row=row_idx, column=2, value=kun['yechilgan'])
                ws3.cell(row=row_idx, column=3, value=kun['togri'])
                ws3.cell(row=row_idx, column=4, value=round(kun['aniqlik'], 1))
        
        # Column widths
        for ws in wb.worksheets:
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15
        
        # Excel saqlash
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    async def _rivojlanish_malumotlari_olish(
        self,
        foydalanuvchi_id: UUID,
        boshlangich_sana: datetime = None,
        tugash_sana: datetime = None
    ) -> Dict[str, Any]:
        """Rivojlanish ma'lumotlarini bazadan oladi."""
        
        # Foydalanuvchi
        from sqlalchemy.orm import selectinload
        foyd_sorov = select(Foydalanuvchi).where(
            Foydalanuvchi.id == foydalanuvchi_id
        ).options(selectinload(Foydalanuvchi.rivojlanish))
        
        foyd_natija = await self.db.execute(foyd_sorov)
        foydalanuvchi = foyd_natija.scalar_one_or_none()
        
        if not foydalanuvchi:
            return {}
        
        rivojlanish = foydalanuvchi.rivojlanish
        
        # Davr
        davr = "Barcha vaqt"
        if boshlangich_sana and tugash_sana:
            davr = f"{boshlangich_sana.strftime('%Y-%m-%d')} - {tugash_sana.strftime('%Y-%m-%d')}"
        
        # Statistika
        statistika = {}
        if rivojlanish:
            jami = rivojlanish.jami_yechilgan or 0
            togri = rivojlanish.togri_javoblar or 0
            statistika = {
                'jami_yechilgan': jami,
                'togri_javoblar': togri,
                'aniqlik': (togri / jami * 100) if jami > 0 else 0,
                'joriy_streak': rivojlanish.joriy_streak or 0,
                'maksimal_streak': rivojlanish.maksimal_streak or 0,
                'jami_ball': rivojlanish.jami_ball or 0,
                'daraja': rivojlanish.daraja or 1,
            }
        
        return {
            'foydalanuvchi': {
                'id': str(foydalanuvchi.id),
                'toliq_ism': foydalanuvchi.toliq_ism,
                'email': foydalanuvchi.email
            },
            'davr': davr,
            'statistika': statistika,
            'kategoriya_statistika': [],  # TODO: kategoriya bo'yicha statistika
            'kunlik_statistika': []  # TODO: kunlik statistika
        }
